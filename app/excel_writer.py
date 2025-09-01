from __future__ import annotations

from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _apply_table_borders(ws, start_row: int, start_col: int, num_rows: int, num_cols: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, start_row + num_rows):
        for c in range(start_col, start_col + num_cols):
            ws.cell(row=r, column=c).border = border


def _next_row(ws) -> int:
    return ws.max_row + 1 if ws.max_row > 0 else 1


def write_pages_to_excel(pages: Dict[int, Dict[str, Any]], output_path: str) -> None:
    wb = Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for page_number in sorted(pages.keys()):
        content = pages[page_number]
        ws = wb.create_sheet(title=f"Page {page_number}")

        # Set some default column widths
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 28

        # Headings
        headings: List[Dict[str, Any]] = content.get("headings") or []
        for heading in headings:
            level = int(heading.get("level", 1))
            text = str(heading.get("text", "")).strip()
            if not text:
                continue
            row = _next_row(ws)
            cell = ws.cell(row=row, column=1, value=text)
            size = 16 - (level - 1) * 2
            if size < 10:
                size = 10
            cell.font = Font(bold=True, size=size)
            cell.alignment = Alignment(wrap_text=True)
            # Visual spacing after headings
            ws.row_dimensions[row].height = 20
            _ = ws.cell(row=row + 1, column=1, value=None)

        # Paragraphs
        paragraphs: List[str] = content.get("paragraphs") or []
        for paragraph in paragraphs:
            text = str(paragraph).strip()
            if not text:
                continue
            row = _next_row(ws)
            cell = ws.cell(row=row, column=1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 40

        # Tables
        tables = content.get("tables") or []
        for table in tables:
            caption = str(table.get("caption", "")).strip()
            headers = [str(h) for h in (table.get("headers") or [])]
            rows = table.get("rows") or []

            if caption:
                row = _next_row(ws)
                cap_cell = ws.cell(row=row, column=1, value=caption)
                cap_cell.font = Font(italic=True, color="555555")
                ws.row_dimensions[row].height = 18

            if headers:
                row = _next_row(ws)
                for idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
                start_row = row
                num_cols = len(headers)
            else:
                start_row = _next_row(ws)
                num_cols = max((len(r) for r in rows), default=1)

            # Data rows
            for r_idx, row_values in enumerate(rows, start=1):
                row = start_row + r_idx
                for c_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row, column=c_idx, value=str(value))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Add borders
            total_rows = (1 if headers else 0) + len(rows)
            if total_rows > 0:
                _apply_table_borders(ws, start_row=(start_row if headers else start_row), start_col=1, num_rows=total_rows, num_cols=num_cols)

            # Spacing after table
            _ = ws.cell(row=_next_row(ws) + 1, column=1, value=None)

    wb.save(output_path)